"""
Simple Excel Character Sheet Analyzer for RPG System
Extracts only essential data needed for system expansion
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# Setup basic logging
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
# If you want to use a custom logger, you can set it up here


class SimpleCharacterSheetAnalyzer:
    """Simple analyzer focused on RPG system essentials"""
    
    def __init__(self, excel_path: str = "char_sheet.xlsx"):
        self.excel_path = Path(__file__).parent / excel_path
        self.analysis = {}
        
    def analyze(self) -> dict:
        """Main analysis method - extracts all essential data"""
        print(f"Analyzing: {self.excel_path}")
        
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        
        # Load workbooks
        wb_formulas = load_workbook(self.excel_path, data_only=False)
        wb_values = load_workbook(self.excel_path, data_only=True)
        
        try:
            self.analysis = {
                "file_info": {
                    "path": str(self.excel_path),
                    "size_kb": round(self.excel_path.stat().st_size / 1024, 1),
                    "sheets": wb_formulas.sheetnames
                },
                "template_data": self._extract_template_essentials(wb_formulas, wb_values),
                "vl_table": self._extract_vl_table(wb_values),
                "validation_rules": self._extract_validation_basics(wb_formulas)
            }
            
            return self.analysis
            
        finally:
            wb_formulas.close()
            wb_values.close()

    def _extract_vl_table(self, wb_values) -> dict:
        """Extract lookup table data"""
        if "VL Table" not in wb_values.sheetnames:
            return {}
            
        sheet = wb_values["VL Table"]
        lookup_data = {}
        
        # Scan columns B and C (typical lookup table format)
        for row in range(2, 100):  # Reasonable range
            key_cell = sheet[f"B{row}"]
            value_cell = sheet[f"C{row}"]
            
            if key_cell.value and value_cell.value:
                lookup_data[str(key_cell.value)] = str(value_cell.value)
        
        return lookup_data
    
    def _extract_validation_basics(self, wb_formulas) -> dict:
        """Extract basic validation info"""
        validation_info = {}
        
        for sheet_name in wb_formulas.sheetnames:
            sheet = wb_formulas[sheet_name]
            
            if sheet.data_validations.dataValidation:
                validation_info[sheet_name] = []
                
                for dv in sheet.data_validations.dataValidation:
                    if dv.formula1:  # Only if there's actual validation
                        validation_info[sheet_name].append({
                            "ranges": [str(r) for r in dv.ranges],
                            "type": dv.type,
                            "formula": dv.formula1
                        })
        
        return validation_info
    
    def _get_bg_color(self, cell) -> str:
        """Get background color as hex string"""
        try:
            if hasattr(cell.fill, 'start_color') and hasattr(cell.fill.start_color, 'rgb'):
                return str(cell.fill.start_color.rgb)
        except:
            pass
        return ""
    
    def _is_bold(self, cell) -> bool:
        """Check if cell text is bold"""
        try:
            return cell.font.bold is True
        except:
            return False
    
    def save_results(self, output_file: str = "analysis_results.json"):
        """Save analysis results to JSON file"""
        output_path = Path(__file__).parent / output_file
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.analysis, f, indent=2, default=str)
        
        print(f"Results saved to: {output_path}")
        print(f"Output size: {round(output_path.stat().st_size / 1024, 1)} KB")
        
        return output_path
    
    def _is_section_header(self, cell) -> bool:
        """Check if cell appears to be a section header"""
        try:
            return (self._is_bold(cell) and 
                    isinstance(cell.value, str) and 
                    len(cell.value) > 3 and
                    not cell.value.startswith('='))
        except:
            return False
    
    def _analyze_input_density(self, input_schema) -> dict:
        """Analyze how many inputs are in each section"""
        density = {}
        for cell_ref, cell_data in input_schema.items():
            section = cell_data["section"]
            if section not in density:
                density[section] = 0
            density[section] += 1
        return density

    def _get_style_name(self, cell) -> str:
        """Get Excel style name instead of hex color"""
        try:
            if hasattr(cell, 'style') and cell.style:
                return str(cell.style).lower()
        except:
            pass
        return ""


    def print_summary(self):
        """Print a summary of what was found"""
        if not self.analysis:
            print("No analysis data available")
            return
        
        print("\n" + "="*50)
        print("ANALYSIS SUMMARY")
        print("="*50)
        
        file_info = self.analysis.get("file_info", {})
        print(f"File: {file_info.get('path', 'Unknown')}")
        print(f"Size: {file_info.get('size_kb', 0)} KB")
        print(f"Sheets: {', '.join(file_info.get('sheets', []))}")
        
        template = self.analysis.get("template_data", {})
        analysis = template.get("analysis", {})
        
        print(f"\nTemplate Sheet:")
        print(f"  Player Inputs: {analysis.get('total_number_inputs', 0)} (number)")
        print(f"  Text Inputs: {analysis.get('total_text_inputs', 0)} (text)")
        print(f"  Dropdowns: {analysis.get('total_dropdowns', 0)}")
        print(f"  Calculations: {analysis.get('total_calculations', 0)}")
        print(f"  Labels: {len(template.get('labels', {}))}")
        
        vl_table = self.analysis.get("vl_table", {})
        print(f"  VL Table Entries: {len(vl_table)}")
        
        validation = self.analysis.get("validation_rules", {})
        total_validations = sum(len(rules) for rules in validation.values())
        print(f"  Validation Rules: {total_validations}")
        
        # Debug style detection
        style_analysis = template.get("style_analysis", {})
        if style_analysis:
            print(f"\nSTYLE DEBUGGING:")
            styles_found = style_analysis.get("styles_found", {})
            print(f"  Total unique styles found: {len(styles_found)}")
            for style_name, count in list(styles_found.items())[:10]:  # Show first 10
                print(f"    {style_name}: {count} cells")
        
        # Show some sample data (with None handling)
        inputs = template.get("input_schema", {})
        if inputs:
            print(f"\nSample Input Cells:")
            for i, (cell_ref, data) in enumerate(list(inputs.items())[:5]):
                value = data.get('value', 'Empty')
                input_type = data.get('input_type', 'unknown')
                style = data.get('style_name', 'no style')
                print(f"  {cell_ref}: {value} ({input_type}) [{style}]")
        
        calculations = template.get("calculations", {})
        if calculations:
            print(f"\nSample Formula Cells:")
            for i, (cell_ref, data) in enumerate(list(calculations.items())[:3]):
                formula = data.get('formula') or ''  # Handle None
                if len(formula) > 50:
                    formula = formula[:50] + "..."
                result = data.get('result', 'N/A')
                print(f"  {cell_ref}: {formula} = {result}")
        
        if vl_table:
            print(f"\nSample VL Table Entries:")
            for i, (key, value) in enumerate(list(vl_table.items())[:3]):
                value_short = value[:50] + "..." if len(value) > 50 else value
                print(f"  {key}: {value_short}")

    def _resolve_data_validation(self, wb_formulas, wb_values) -> dict:
        """Resolve named ranges in data validation to actual dropdown options"""
        resolved = {}
        
        # Get validation rules
        template_sheet = wb_formulas["Template"]
        if not template_sheet.data_validations.dataValidation:
            return resolved
        
        for dv in template_sheet.data_validations.dataValidation:
            if dv.formula1 and dv.type == "list":
                range_refs = [str(r) for r in dv.ranges]
                named_range = dv.formula1
                
                # Try to resolve named range to actual values
                actual_options = self._resolve_named_range(wb_formulas, wb_values, named_range)
                
                resolved[named_range] = {
                    "applies_to_cells": range_refs,
                    "options": actual_options
                }
        
        return resolved

    def _resolve_named_range(self, wb_formulas, wb_values, named_range: str) -> list:
        """Resolve a named range like 'ArcAT' to actual dropdown options"""
        try:
            # Check if it's a simple list first
            if named_range.startswith('"') and named_range.endswith('"'):
                # It's a literal list like "Resolve,Stability,Vitality"
                content = named_range.strip('"')
                return content.split(',')
            
            # Try to find it as a named range
            if hasattr(wb_formulas, 'defined_names') and named_range in wb_formulas.defined_names:
                # This is a proper named range - would need more complex resolution
                return [f"Named range: {named_range}"]
            
            # Check if it matches VL table columns or patterns
            vl_sheet = wb_values.get("VL Table")
            if vl_sheet:
                # Look for column headers that might match
                for col in range(1, 10):
                    cell_val = vl_sheet[f"{get_column_letter(col)}1"].value
                    if cell_val and str(cell_val).replace(" ", "") == named_range:
                        # Found matching column, extract options
                        options = []
                        for row in range(2, 100):
                            val = vl_sheet[f"{get_column_letter(col)}{row}"].value
                            if val:
                                options.append(str(val))
                            else:
                                break
                        return options
            
            return [f"Unresolved: {named_range}"]
            
        except Exception as e:
            logger.debug(f"Could not resolve named range {named_range}: {e}")
            return [f"Error resolving: {named_range}"]


    def _extract_template_essentials(self, wb_formulas, wb_values) -> dict:
        """Extract template STRUCTURE - FIXED formula extraction"""
        
        # Check which sheet to analyze - prefer Pandora if it exists
        target_sheet = None
        if "Pandora" in wb_formulas.sheetnames:
            target_sheet = "Pandora"
            logger.info("Using Pandora sheet for analysis")
        elif "Template" in wb_formulas.sheetnames:
            target_sheet = "Template"
            logger.info("Using Template sheet for analysis")
        else:
            logger.error("Neither Pandora nor Template sheet found")
            return {}
            
        sheet_f = wb_formulas[target_sheet]
        sheet_v = wb_values[target_sheet]
        
        data = {
            "input_schema": {},      
            "calculations": {},      
            "labels": {},           
            "sections": {},         
            "formula_debug": {},     # Debug info
            "validation_resolved": {}
        }
        
        current_section = "unknown"
        formulas_found = 0
        formulas_missed = 0
        
        # Scan ALL cells with BETTER formula detection
        for row in range(2, 241):
            for col in range(1, 9):
                cell_ref = f"{get_column_letter(col)}{row}"
                cell_f = sheet_f[cell_ref]
                cell_v = sheet_v[cell_ref]
                
                # Get style and color
                style_name = self._get_style_name(cell_f)
                bg_color = self._get_bg_color(cell_f)
                
                # COMPREHENSIVE formula extraction
                formula_text = self._extract_formula_comprehensive(cell_f, cell_ref)
                
                # Debug specific cell the user mentioned
                if cell_ref == "E192":
                    data["formula_debug"]["E192"] = {
                        "cell_value": cell_f.value,
                        "cell_data_type": getattr(cell_f, 'data_type', 'unknown'),
                        "has_formula_attr": hasattr(cell_f, 'formula'),
                        "formula_attr": getattr(cell_f, 'formula', None),
                        "extracted_formula": formula_text,
                        "style": style_name
                    }
                
                # Classify cell type
                if style_name and any(s in style_name.lower() for s in ["good", "input", "note"]):
                    input_type = self._classify_input_type(style_name)
                    data["input_schema"][cell_ref] = {
                        "input_type": input_type,
                        "style_name": style_name,
                        "value": cell_v.value
                    }
                    
                elif style_name and "calculation" in style_name.lower():
                    # Store calculation with proper formula
                    data["calculations"][cell_ref] = {
                        "formula": formula_text,
                        "result": cell_v.value,
                        "style_name": style_name,
                        "section": current_section
                    }
                    
                    if formula_text:
                        formulas_found += 1
                    else:
                        formulas_missed += 1
                
                # Track section headers
                elif cell_f.value and isinstance(cell_f.value, str) and not str(cell_f.value).startswith('='):
                    if self._is_section_header(cell_f):
                        current_section = cell_f.value
                        data["sections"][current_section] = {"start_row": row}
                    
                    data["labels"][cell_ref] = cell_f.value
        
        logger.info(f"Formula extraction: {formulas_found} found, {formulas_missed} missed")
        
        # Analysis summary  
        data["analysis"] = {
            "sheet_analyzed": target_sheet,
            "total_number_inputs": len([x for x in data["input_schema"].values() if "number" in x["input_type"]]),
            "total_text_inputs": len([x for x in data["input_schema"].values() if "text" in x["input_type"]]),
            "total_dropdowns": len([x for x in data["input_schema"].values() if "dropdown" in x["input_type"]]),
            "total_calculations": len(data["calculations"]),
            "formulas_found": formulas_found,
            "formulas_missed": formulas_missed,
            "sections_found": list(data["sections"].keys()),
        }
        
        return data

    def _extract_formula_comprehensive(self, cell, cell_ref: str) -> str:
        """Try multiple methods to extract formula from cell - FIXED for ArrayFormulas"""
        try:
            # Method 1: Check cell.value for formula string
            if cell.value is not None:
                val_str = str(cell.value)
                if val_str.startswith('='):
                    return val_str
            
            # Method 2: Handle ArrayFormula objects specifically
            if hasattr(cell, 'value') and cell.value is not None:
                cell_val = cell.value
                # Check if it's an ArrayFormula object
                if hasattr(cell_val, 'text'):
                    formula_text = str(cell_val.text)
                    if formula_text and not formula_text == 'None':
                        return f"={formula_text}" if not formula_text.startswith('=') else formula_text
                elif hasattr(cell_val, '__class__') and 'ArrayFormula' in str(cell_val.__class__):
                    # It's an ArrayFormula object, try to extract the formula
                    if hasattr(cell_val, 'ref') and hasattr(cell_val, 'text'):
                        return f"={cell_val.text}"
                    # Try other ArrayFormula attributes
                    for attr in ['formula', 'text', '_text', 'value']:
                        if hasattr(cell_val, attr):
                            attr_val = getattr(cell_val, attr)
                            if attr_val and str(attr_val) not in ['None', '']:
                                formula_text = str(attr_val)
                                return f"={formula_text}" if not formula_text.startswith('=') else formula_text
            
            # Method 3: Check formula attribute
            if hasattr(cell, 'formula') and cell.formula:
                formula_str = str(cell.formula)
                if formula_str and not formula_str == 'None':
                    return formula_str
            
            # Method 4: Check _value attribute (internal openpyxl)
            if hasattr(cell, '_value') and cell._value is not None:
                val_str = str(cell._value)
                if val_str.startswith('='):
                    return val_str
            
            # Method 5: Direct access to worksheet formulas (last resort)
            try:
                worksheet = cell.parent
                if hasattr(worksheet, '_formulas') and cell.coordinate in worksheet._formulas:
                    return worksheet._formulas[cell.coordinate]
            except:
                pass
            
            # Method 6: Check if it's a formula cell by data type but couldn't extract
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                # It's a formula cell but we couldn't extract it
                logger.warning(f"Formula cell {cell_ref} detected but formula not extracted - data type: {cell.data_type}")
                return f"FORMULA_PRESENT_BUT_NOT_EXTRACTED_{cell_ref}"
                
            return None
            
        except Exception as e:
            logger.debug(f"Error extracting formula from {cell_ref}: {e}")
            return None


    def _classify_input_type(self, style_name: str) -> str:
        """Classify input type from style name"""
        style_lower = style_name.lower()
        if "good" in style_lower:
            return "text_input"
        elif "input" in style_lower:
            return "number_input"  
        elif "note" in style_lower:
            return "dropdown_input"
        else:
            return "unknown_input"


def main():
    """Main function to run the analysis"""
    try:
        # Create analyzer
        analyzer = SimpleCharacterSheetAnalyzer()
        
        # Run analysis
        results = analyzer.analyze()
        
        # Print summary
        analyzer.print_summary()
        
        # Save results
        output_file = analyzer.save_results()
        
        # Calculate size comparison
        original_size = analyzer.analysis["file_info"]["size_kb"]
        output_size = round(output_file.stat().st_size / 1024, 1)
        
        print(f"\n" + "="*50)
        print("SIZE COMPARISON")
        print("="*50)
        print(f"Original Excel: {original_size} KB")
        print(f"Extracted JSON: {output_size} KB")
        
        if output_size > original_size:
            print(f"❌ Size INCREASED by {round((output_size/original_size - 1) * 100, 1)}%")
        else:
            print(f"✅ Size REDUCED by {round((1 - output_size/original_size) * 100, 1)}%")
        
        print(f"\n✅ Analysis complete!")
        
    except Exception as e:
        print(f"❌ Analysis failed: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()